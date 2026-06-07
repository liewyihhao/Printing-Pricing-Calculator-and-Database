"""Phase 2 step 1 — capture ONE Loose Sheet combo two ways and compare:
  (A) Generate Price List -> scrape the on-page table
  (B) Export to Excel    -> save + parse the downloaded file

Pick a stable, common combo: A4 size + Gloss Art Card 250gsm (a "Best Seller").
Run: python compare_capture.py
"""
import os, sys, json, asyncio
from pathlib import Path
from playwright.async_api import async_playwright

from excard import log, polite_pause, launch_browser, login, OUT

PRICE_URL = os.getenv("EXCARD_PRICE_URL", "https://www.excard.com.my/price-list/Litho/21").strip()

# Selectors discovered on the Loose Sheet matrix.
SEL = {
    "size": "select[name='ctl00$mainContent$order_spec_controller1$order_spec_standard_matrix1$ddlSize']",
    "paper": "select[name='ctl00$mainContent$order_spec_controller1$order_spec_standard_matrix1$ddlPaper']",
}
SIZE_LABEL = "A4 (210mm x 297mm)"
PAPER_LABEL = "Gloss Art Card 250gsm (2 sides coated) - Best Seller"
# Lamination/Finishing appears dynamically after Paper and is REQUIRED.
LAMINATION_LABEL = "Matte Lamination (Front)"


async def select_dynamic_by_option(page, option_label):
    """Select `option_label` in whichever <select> currently contains it
    (used for fields like Lamination that are injected after another choice)."""
    sel = page.locator("select").filter(
        has=page.locator(f"option:text-is('{option_label}')")).first
    if await sel.count() == 0:
        log(f"⚠ No <select> currently offers '{option_label}'.")
        return False
    name = await sel.get_attribute("name")
    log(f"Selecting '{option_label}' (field: {name})")
    await sel.select_option(label=option_label)
    try:
        await page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    await polite_pause()
    return True


async def click_by_text(page, text):
    """Find Generate/Export controls regardless of whether they're button/a/input."""
    for sel in [
        f"input[type='submit'][value*='{text}' i]",
        f"button:has-text('{text}')",
        f"a:has-text('{text}')",
        f"input[type='button'][value*='{text}' i]",
        f"text={text}",
    ]:
        loc = page.locator(sel).first
        try:
            if await loc.count() > 0:
                return loc
        except Exception:
            continue
    return None


async def select_option_safe(page, selector, label):
    """Select by visible label; ASP.NET autopostback may reload — wait it out."""
    log(f"Selecting '{label}'")
    await page.select_option(selector, label=label)
    try:
        await page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    await polite_pause()


async def scrape_tables(page):
    """Return every HTML table on the page as list-of-rows, plus which look like
    a price/quantity table (contain digits + 'RM' or 'Qty')."""
    return await page.evaluate(
        """() => {
            const tables = [...document.querySelectorAll('table')].map(t => {
                const rows = [...t.querySelectorAll('tr')].map(tr =>
                    [...tr.querySelectorAll('th,td')].map(c => c.innerText.trim()));
                const flat = rows.flat().join(' ');
                // A real price table has RM amounts AND no giant option-list cells.
                const hasRM = /RM\\s*\\d|\\d+\\.\\d{2}/.test(flat);
                const maxCell = Math.max(0, ...rows.flat().map(c => c.length));
                const looksPricey = hasRM && maxCell < 120;
                return { rows, looksPricey, rowCount: rows.length, maxCell };
            });
            return tables.filter(t => t.rowCount > 1);
        }"""
    )


async def main():
    async with async_playwright() as pw:
        browser = await launch_browser(pw)
        ctx = await browser.new_context(accept_downloads=True,
                                        viewport={"width": 1440, "height": 1000})
        page = await ctx.new_page()
        try:
            if not await login(page):
                raise SystemExit("Login failed; cannot continue.")

            log(f"Opening {PRICE_URL}")
            await page.goto(PRICE_URL, wait_until="domcontentloaded")
            await polite_pause()

            await select_option_safe(page, SEL["size"], SIZE_LABEL)
            await select_option_safe(page, SEL["paper"], PAPER_LABEL)
            # Lamination/Finishing is injected after Paper and is required.
            await select_dynamic_by_option(page, LAMINATION_LABEL)

            # Quantity is dynamic (populated from hDynamicQty, e.g. "250,500")
            # and required for a price to compute. Pick the first available qty.
            qtys = await page.evaluate(
                """() => {
                    const h = document.querySelector("input[id*='hDynamicQty']");
                    return h ? h.value.split(',').filter(Boolean) : [];
                }""")
            log(f"Available quantities: {qtys}")
            if qtys:
                qsel = page.locator("select").filter(
                    has=page.locator(f"option:text-is('{qtys[0]}')")).first
                if await qsel.count():
                    log(f"Selecting quantity {qtys[0]} (field: {await qsel.get_attribute('name')})")
                    await qsel.select_option(label=qtys[0])
                    try:
                        await page.wait_for_load_state("networkidle", timeout=15000)
                    except Exception:
                        pass
                    await polite_pause()
                else:
                    log("⚠ Could not locate the quantity <select>.")
            # Save the fully-configured form HTML so we capture dynamic field names.
            (OUT / "compare_form_configured.html").write_text(
                await page.content(), encoding="utf-8")

            # ---- (A) Generate Price List -> scrape ----
            gen = page.locator("#mainContent_btnPriceList")
            if await gen.count() == 0:
                gen = await click_by_text(page, "Generate Price List")
            if gen:
                log("Clicking 'Generate Price List' (AJAX postback)")
                # It's an ASP.NET AJAX __doPostBack — don't wait for navigation.
                await gen.click(no_wait_after=True)
                # Wait for an RM amount to appear anywhere on the page.
                try:
                    await page.wait_for_function(
                        "() => /RM\\s*\\d|\\d+\\.\\d{2}/.test(document.body.innerText)",
                        timeout=25000)
                    log("Price content detected.")
                except Exception:
                    log("⚠ No RM price content appeared within 25s (validation may have blocked it).")
                await polite_pause()
            else:
                log("⚠ Could not find 'Generate Price List' button.")

            await page.screenshot(path=str(OUT / "compare_generated.png"), full_page=True)
            tables = await scrape_tables(page)
            pricey = [t for t in tables if t["looksPricey"]]
            (OUT / "compare_scraped_tables.json").write_text(
                json.dumps(tables, indent=2, ensure_ascii=False), encoding="utf-8")
            log(f"(A) Scraped {len(tables)} tables, {len(pricey)} look price-like.")
            for i, t in enumerate(pricey[:2]):
                log(f"    -- price table #{i}: {t['rowCount']} rows --")
                for r in t["rows"][:6]:
                    log(f"      {[c[:18] for c in r][:10]}")

            # ---- (B) Export to Excel -> download ----
            exp = await click_by_text(page, "Export to Excel")
            if exp:
                log("Clicking 'Export to Excel'")
                try:
                    async with page.expect_download(timeout=20000) as dl_info:
                        await exp.click()
                    dl = await dl_info.value
                    suggested = dl.suggested_filename or "export.bin"
                    dest = OUT / f"compare_export_{suggested}"
                    await dl.save_as(str(dest))
                    log(f"(B) Downloaded export -> {dest.name} ({dest.stat().st_size} bytes)")
                    parse_export(dest)
                except Exception as e:
                    log(f"(B) Export download did not fire as a file: {e!r}")
                    log("    (It may render inline instead — captured in scraped tables.)")
            else:
                log("⚠ Could not find 'Export to Excel' button.")

            log("DONE. Review crawler/output/compare_* files for the side-by-side.")
        finally:
            await polite_pause()
            await browser.close()


def parse_export(path: Path):
    """Best-effort parse of the exported file (xlsx/xls/csv/html-disguised-as-xls)."""
    head = path.read_bytes()[:8]
    log(f"    export magic bytes: {head!r}")
    ext = path.suffix.lower()
    try:
        if head[:2] == b"PK" or ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = [[c for c in row] for row in ws.iter_rows(values_only=True)][:8]
        elif head[:4] == b"\xd0\xcf\x11\xe0" or ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(path); ws = wb.sheet_by_index(0)
            rows = [ws.row_values(r) for r in range(min(8, ws.nrows))]
        else:
            # Excard often exports an HTML table with a .xls name.
            text = path.read_text(encoding="utf-8", errors="replace")
            if "<table" in text.lower():
                log("    export is HTML-as-xls; parse with same table logic later.")
                return
            rows = [ln.split(",") for ln in text.splitlines()[:8]]
        log("    first rows of export:")
        for r in rows:
            log(f"      {r}")
    except Exception as e:
        log(f"    parse error: {e!r}")


if __name__ == "__main__":
    asyncio.run(main())
